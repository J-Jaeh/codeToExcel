import sys
import os
from bs4 import BeautifulSoup
import openpyxl as oxl
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import Tk
from tkinter import messagebox as msg
import shutil



def find_tables_with_methods(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')
    a_tags = [table.find('a', {'id': lambda x: x and '-methods' in x}) for table in tables]

    # Filter out None values (a_tags that don't meet the condition)
    filtered_tables = [table for table, a_tag in zip(tables, a_tags) if a_tag]

    return filtered_tables


def find_title_text(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    title_tag = soup.find('title')
    if title_tag:
        return title_tag.get_text()
    else:
        return ""
    
def find_source_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()
        soup = BeautifulSoup(html_content, 'html.parser')

        # "a" 태그의 클래스 이름이 "el"인 모든 "a" 태그를 찾습니다.
        el_a_tags = soup.find_all('a', class_='el')
        
        if el_a_tags:
            # 가장 마지막에 나오는 "a" 태그의 내용을 반환합니다.
            last_el_a_tag = el_a_tags[-1]
            return last_el_a_tag.text.strip()
        
        # "a" 태그의 클래스 이름이 "el"인 태그가 없는 경우 None을 반환합니다.
        return ""
def toCharParamter(prototype):
    orginalPrototype = prototype.replace("정적",'').strip() # 정적 ~ 한글 제거
    allParameterValue=re.search(r'\(\s*(.*?)\s*\)', orginalPrototype).group(1)
    parameterInMap = list(allParameterValue)
    saveParameter =[]   # <> 인식하는 컴마로 구분된 파라미터데이터타입 + 파라미터 이름이 모인 리스트다 ; for 문 돟ㄹ려보자
    saveStack=[]
    para=''
    for str in parameterInMap:
        if str=="<":
            saveStack.append(str)
            para+=str
            continue
        elif len(saveStack)!=0:
            if str == '>':
                saveStack.pop()
                para+=str
            else:
                para+=str
            continue               
        elif not saveStack:       
            if str !=",":
                para+=str
            else:
                saveParameter.append(para)    
                para=''
    saveParameter.append(para)            
    return saveParameter
    
def create_table_DDS(ws, start_row,prototype,parameterColumCorrectionValue):
    #-----------------------------------컬럼 보정작업을 생각해보자---------
    orginalPrototype = prototype.replace("정적",'').strip() # 정적 ~ 한글 제거

    dividedPrototype =orginalPrototype.split('(') # "(" 기준으로 split

    unitNameList=dividedPrototype[0].strip().split(' ')  # 접근제어자 ,(static), void , 메소드이름  이러면 차라리 나누기전 문자가 있냐 ? orginalPrototype에서 void 있으면 

    unitNameIndex=len(unitNameList)
    allParameterValue=re.search(r'\(\s*(.*?)\s*\)', orginalPrototype).group(1)
    parameterList = allParameterValue.split(",")
    
    ws.insert_rows(start_row+4,parameterColumCorrectionValue)
    #----------------------------------------------------------------------

    #------------------------------------기본 테이블 설정값이다!--------------

    ws[f'A{start_row-1}'].font = Font(name='Arial', size=20, bold=True)
    ws[f'B{start_row}'] = 'Software Unit Information'
    ws[f'B{start_row + 1}'] = 'Unit ID'
    ws[f'D{start_row + 1}'] = 'Unit Name'
    ws[f'F{start_row + 1}'] = 'ASIL'
    ws[f'G{start_row + 1}'] = 'QM'
    ws[f'B{start_row + 2}'] = 'Prototype'
    ws[f'B{start_row + 3}'] = 'Parameter'
    ws[f'C{start_row + 3}'] = 'Data Type'
    ws[f'D{start_row + 3}'] = 'Name'
    ws[f'E{start_row + 3}'] = 'Range'
    ws[f'F{start_row + 3}'] = 'IN/OUT'
    ws[f'G{start_row + 3}'] = 'Description'
 

    #------------컬럼 보정작업을 하고 병합을하고 병합한 셀에 값을 추가해줘야 오류가 없다 !-----

    ws.merge_cells(f'B{start_row}:G{start_row}') #Software Unit Information
    ws.merge_cells(f'C{start_row+2}:G{start_row+2}') # 프로토타입 들어갈 공간 즉 파싱 값이 들어갈 자리
    ws.merge_cells(f'B{start_row+3}:B{start_row+4+parameterColumCorrectionValue}') #Parameter
    ws.merge_cells(f'B{start_row+parameterColumCorrectionValue+5}:B{start_row+parameterColumCorrectionValue+6}') #Return Value
    ws.merge_cells(f'D{start_row+parameterColumCorrectionValue+5}:E{start_row+parameterColumCorrectionValue+5}') #Possible Return Value
    ws.merge_cells(f'F{start_row+parameterColumCorrectionValue+5}:G{start_row+parameterColumCorrectionValue+5}') #Description
    ws.merge_cells(f'D{start_row+parameterColumCorrectionValue+6}:E{start_row+parameterColumCorrectionValue+6}') #Possible Return Value 값이 들어갈 공간
    ws.merge_cells(f'F{start_row+6+parameterColumCorrectionValue}:G{start_row+6+parameterColumCorrectionValue}') #Description에 대한 설명이 들어갈 공간
    ws.merge_cells(f'B{start_row+7+parameterColumCorrectionValue}:B{start_row+8+parameterColumCorrectionValue}') #Imported Class or Global Value
    ws.merge_cells(f'C{start_row+9+parameterColumCorrectionValue}:G{start_row+9+parameterColumCorrectionValue}') #메소드에 대한 Description에 대한 내용이 들어간 공간
    ws.merge_cells(f'C{start_row+10+parameterColumCorrectionValue}:G{start_row+10+parameterColumCorrectionValue}') #Called Function 내용이 들어갈 공간
    ws.merge_cells(f'C{start_row+11+parameterColumCorrectionValue}:G{start_row+11+parameterColumCorrectionValue}') #calling Function 내용이 들어갈 공간
    ws.merge_cells(f'C{start_row+12+parameterColumCorrectionValue}:G{start_row+12+parameterColumCorrectionValue}')


    ws[f'B{start_row + 5+parameterColumCorrectionValue}'] = 'Return Value'
    ws[f'C{start_row + 5+parameterColumCorrectionValue}'] = 'Data Type'
    ws[f'D{start_row + 5+parameterColumCorrectionValue}'] = 'Possible Return Value'
    ws[f'F{start_row + 5+parameterColumCorrectionValue}'] = 'Description'
    
    ws[f'B{start_row + 7+parameterColumCorrectionValue}'] = 'Imported Class or Global Value'
    ws[f'C{start_row + 7+parameterColumCorrectionValue}'] = 'Data Type'
    ws[f'D{start_row + 7+parameterColumCorrectionValue}'] = 'Name'
    ws[f'E{start_row + 7+parameterColumCorrectionValue}'] = 'Range'
    ws[f'F{start_row + 7+parameterColumCorrectionValue}'] = 'Read/Write'
    ws[f'G{start_row + 7+parameterColumCorrectionValue}'] = 'Description'
    ws[f'F{start_row + 8+parameterColumCorrectionValue}'] = '-' # 이건 배경색 안들어감  Read/Write를 적는칸
    ws[f'B{start_row + 9+parameterColumCorrectionValue}'] = 'Description'
    ws[f'B{start_row + 10+parameterColumCorrectionValue}'] = 'Called Function'
    ws[f'B{start_row + 11+parameterColumCorrectionValue}'] = 'Calling Function'
    ws[f'B{start_row + 12+parameterColumCorrectionValue}'] = 'Dynamic View Design\r(Control Flow Diagram)'
    #---------------------------------------------------------------------------

    #---------------------파싱값 할당--------------------------------------------
    ws[f'E{start_row+1}']= unitNameList[unitNameIndex-1]  #unitName 이다 후... 여기를 바꿔야함 ;;; 미쳤다리 ,,,,,,,,,



    ws[f'C{start_row+2}']= orginalPrototype   #prototype 이다 !

    if ">" in allParameterValue:
        saveParameter = toCharParamter(prototype=prototype)   
        for index in range(0,len(saveParameter)):  # 저장된 파라미터 갯수만큼 돌아가지 
            parameterDataType=''
            for combindParameterDataTypeIndex in range(0,len(saveParameter[index].split(" "))-1): # 목적 파라미터 이름 빼고 앞에 있는걸 획득하기 위함 ; 보통 0 1 로 끝나겠지
                parameterDataType+=saveParameter[index].split(" ")[combindParameterDataTypeIndex]  # 이렇게 해주면 되자나 
            ws[f'C{start_row+4+index}'] = parameterDataType # 파라미터 데이터 타입
            ws[f'D{start_row+4+index}'] = saveParameter[index].split(" ")[len(saveParameter[index].split(" "))-1] # 파라미터 이름
            ws[f'F{start_row+4+index}'] = "IN"
            ws[f'C{start_row+4+index}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
            ws[f'D{start_row+4+index}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
            ws[f'F{start_row+4+index}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)        
    else :
        for param in parameterList:
            if not param:
                ws[f'C{start_row+4}'] ='-'
                ws[f'D{start_row+4}'] ='-'
                ws[f'E{start_row+4}'] ='-'
                ws[f'F{start_row+4}'] ='-'
                ws[f'G{start_row+4}'] ='-'
            elif len(parameterList) ==1:
                parameter = allParameterValue.split(' ')
                dataType=''
                for index in range(0,len(parameter)-1):
                    dataType+=parameter[index]+" "
                ws[f'C{start_row+4}'] = dataType # 파라미터 데이터 타입
                ws[f'D{start_row+4}'] = parameter[len(parameter)-1] # 파라미터 이름
                ws[f'F{start_row+4}'] = "IN"
            else: # 이제 파라미터 값이 여러개 오는 짜증나는 경우다 // 일단 여러줄 만드는거 성공했자나 ;; 이제 값만 잘 넣어주면된다...
                # for문 이용해서 행값 조절 좀 해줘야 할듯 ?
                for parameterColum in range(0,parameterColumCorrectionValue+1):
                    parameter = parameterList[parameterColum].split(' ')
                    dataType=''
                    for index in range(0,len(parameter)-1):
                        dataType+=parameter[index]+" "
                    ws[f'C{start_row+4+parameterColum}'] = dataType # 파라미터 데이터 타입
                    ws[f'D{start_row+4+parameterColum}'] = parameter[len(parameter)-1] # 파라미터 이름
                    ws[f'F{start_row+4+parameterColum}'] = "IN"
                    ws[f'C{start_row+4+parameterColum}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
                    ws[f'D{start_row+4+parameterColum}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                    ws[f'F{start_row+4+parameterColum}'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 

    # ----------------적절한 반환형을 주입시키는 분기------------------------------
    if unitNameIndex == 2:
        ws[f'C{start_row+6+parameterColumCorrectionValue}'] = "-"   
        ws[f'D{start_row+6+parameterColumCorrectionValue}'] = "-"
        ws[f'F{start_row+6+parameterColumCorrectionValue}'] = "-"
    elif unitNameIndex == 3:
        if unitNameList[unitNameIndex-2] == "void" or unitNameList[unitNameIndex-2] == "virtual" :
            ws[f'C{start_row+6+parameterColumCorrectionValue}'] = "-"  
            ws[f'D{start_row+6+parameterColumCorrectionValue}'] = "-"
            ws[f'F{start_row+6+parameterColumCorrectionValue}'] = "-"
        else:
            ws[f'C{start_row+6+parameterColumCorrectionValue}'] = unitNameList[unitNameIndex-2] # 반환형!
    elif unitNameIndex > 3:
        if "void" in orginalPrototype:
            ws[f'C{start_row+6+parameterColumCorrectionValue}'] = "-"  
            ws[f'D{start_row+6+parameterColumCorrectionValue}'] = "-"
            ws[f'F{start_row+6+parameterColumCorrectionValue}'] = "-"
        else:    
            returnValue=''
            for unitIndex in range(1,unitNameIndex-1):
                returnValue +=unitNameList[unitIndex]+" "
            ws[f'C{start_row+6+parameterColumCorrectionValue}']=returnValue


    #---------------------------------------------------------------------------
    #---------------------------------------------------------------------------


    # 열 크기 지정
    col_widths = {"B": 20, "C": 30, "D": 30, "E": 30, "F": 30, "G": 30}
    for col_name in col_widths:
        ws.column_dimensions[col_name].width = col_widths[col_name]

    ws.row_dimensions[start_row + 12+parameterColumCorrectionValue].height = 400
    # 색 및 글자 폰트
    col_color = [f'B{start_row}', f'B{start_row + 1}', f'D{start_row + 1}', f'F{start_row + 1}', f'B{start_row + 2}', 
                 f'B{start_row + 3}', f'C{start_row + 3}', f'D{start_row + 3}', f'E{start_row + 3}', f'F{start_row + 3}', 
                 f'G{start_row + 3}', f'B{start_row + 5+parameterColumCorrectionValue}', f'C{start_row + 5+parameterColumCorrectionValue}', f'D{start_row + 5+parameterColumCorrectionValue}', f'F{start_row + 5+parameterColumCorrectionValue}', 
                 f'B{start_row + 7+parameterColumCorrectionValue}', f'C{start_row + 7+parameterColumCorrectionValue}', f'D{start_row + 7+parameterColumCorrectionValue}', f'E{start_row + 7+parameterColumCorrectionValue}', f'F{start_row + 7+parameterColumCorrectionValue}', 
                 f'G{start_row + 7+parameterColumCorrectionValue}', f'B{start_row + 9+parameterColumCorrectionValue}', f'B{start_row + 10+parameterColumCorrectionValue}', f'B{start_row + 11+parameterColumCorrectionValue}', f'B{start_row + 12+parameterColumCorrectionValue}']
    for cell in col_color:
        ws[cell].fill = PatternFill(patternType="solid", fgColor='BFBFBF')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        if cell == f'B{start_row}':
            ws[cell].font = Font(name='Arial', size=10, bold=True)
        else:
            ws[cell].font = Font(name='Arial', size=10)
    
    # 추가 정렬작업
    col_center =[f'G{start_row + 1}',f'C{start_row+2}',f'E{start_row+1}',f'C{start_row+4}',f'D{start_row+4}',f'E{start_row+4}',f'F{start_row+4}',f'G{start_row+4}',f'F{start_row + 8+parameterColumCorrectionValue}',f'C{start_row+6+parameterColumCorrectionValue}',f'D{start_row+6+parameterColumCorrectionValue}',f'F{start_row+6+parameterColumCorrectionValue}']
    for cell in col_center:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
 
    #표 테두리 작업
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for rows in ws.iter_rows(min_row=start_row, max_row=start_row+parameterColumCorrectionValue + 12, min_col=2, max_col=7):
        for cell in rows:
            cell.border = border

def create_table_ADS(ws, start_row,prototype,parameterColumCorrectionValue):
    #-----------------------------------컬럼 보정작업을 생각해보자---------
    orginalPrototype = prototype.replace("정적",'').strip() # 정적 ~ 한글 제거

    dividedPrototype =orginalPrototype.split('(') # "(" 기준으로 split

    unitNameList=dividedPrototype[0].strip().split(' ')  # 접근제어자 ,(static), void , 메소드이름  이러면 차라리 나누기전 문자가 있냐 ? orginalPrototype에서 void 있으면 

    unitNameIndex=len(unitNameList)
    allParameterValue=re.search(r'\(\s*(.*?)\s*\)', orginalPrototype).group(1)
    parameterList = allParameterValue.split(",")
    
    ws.insert_rows(start_row+4,parameterColumCorrectionValue)
    #----------------------------------------------------------------------
    # 값 및 범위 할당 
    ClassName_Cell=f'A{start_row-1}'

    InterfaceName_Cell = f'E{start_row}'
    InterfaceName_Value_Cell=f'I{start_row}'
    InterfaceName_MergeCell=InterfaceName_Cell+f':H{start_row}'
    InterfaceName_Value_MergeCell=InterfaceName_Value_Cell+f':AP{start_row}'

    Prototype_Cell = f'E{start_row+1}'
    Prototype_Value_Cell = f'I{start_row+1}'
    Prototype_MergeCell = Prototype_Cell+f':H{start_row+1}'
    Prototype_Value_MegrgeCell=Prototype_Value_Cell+f':AP{start_row+1}'
        
    Description_Cell= f'E{start_row+2}'
    Description_Value_Cell = f'I{start_row+2}'
    Description_MergeCell = Description_Cell+f':H{start_row+2}'
    Description_Value_MergeCell = Description_Value_Cell+f':AP{start_row+2}'

    Parameter_Cell=f'E{start_row+3}' 
    Parameter_MergeCell=Parameter_Cell+f':H{start_row+parameterColumCorrectionValue+4}'

    ParameterDataType_Cell=f'I{start_row+3}'
    ParameterDataType_MergeCell=ParameterDataType_Cell+f':M{start_row+3}'


    #이렇게 좌표로 지정해줘서 변하는 값에 대입해주는게 좋을듯 ㅎㅎ;;
    #파라미터 데이터타입 머지할 좌표 해당row 어차피 한줄이니 상관없음
    Parameter_Element_Value_Merge_Row=start_row+4
    Parameter_Element_Value_Row=Parameter_Element_Value_Merge_Row

    #파라미터 데이터타입 들어갈 열
    ParameterDataType_Value_Culum=9
    ParameterDataType_Value_Merge_Start_Culum=ParameterDataType_Value_Culum
    ParameterDataType_Value_Merge_End_Culum=ParameterDataType_Value_Merge_Start_Culum+4
     # 행 열 행 열 임 !
    #ParameterDataType_Value_Merge_Range=Parameter_Element_Value_Merge_Row,ParameterDataType_Value_Merge_Start_Culum,Parameter_Element_Value_Merge_Row,ParameterDataType_Value_Merge_End_Culum

    ParameterName_Value_Culum=14
    ParameterName_Value_Start_Culum=ParameterName_Value_Culum
    ParameterName_Value_End_Culum=ParameterName_Value_Start_Culum+4
    #ParameterName_Value_Merge_Range=f'({Parameter_Element_Value_Merge_Row},{ParameterName_Value_Start_Culum},{Parameter_Element_Value_Merge_Row},{ParameterName_Value_End_Culum})'

    ParameterRange_Value_Culum=19
    ParameterRange_Value_Start_Culum=ParameterRange_Value_Culum
    ParameterRange_Value_End_Culum=ParameterRange_Value_Start_Culum+4
    #ParameterRange_Value_Merge_Range=f'({Parameter_Element_Value_Merge_Row},{ParameterRange_Value_Start_Culum},{Parameter_Element_Value_Merge_Row},{ParameterRange_Value_End_Culum})'

    ParameterInOrOut_Value_Culum=24
    ParameterInOrOut_Value_Start_Culum=ParameterInOrOut_Value_Culum
    ParameterInOrOut_Value_End_Culum=ParameterInOrOut_Value_Start_Culum+2
    #ParameterInOrOut_Value_Merge_Range=f'({Parameter_Element_Value_Merge_Row},{ParameterInOrOut_Value_Start_Culum},{Parameter_Element_Value_Merge_Row},{ParameterInOrOut_Value_End_Culum})'

    ParameterDescription_Value_Culum=27
    ParameterDescription_Value_Start_Culum=ParameterDescription_Value_Culum
    ParameterDescription_Value_End_Culum=ParameterDescription_Value_Start_Culum+15
    #ParameterDescription_Value_Merge_Range=f'({Parameter_Element_Value_Merge_Row},{ParameterDescription_Value_Start_Culum},{Parameter_Element_Value_Merge_Row},{ParameterDescription_Value_End_Culum})'

    ParameterName_Cell=f'N{start_row+3}'
    ParameterName_MergeCell=ParameterName_Cell+f':R{start_row+3}'

    ParameterRange_Cell=f'S{start_row+3}'
    ParameterRange_MergeCell=ParameterRange_Cell+f':W{start_row+3}'

    ParameterInOrOut_Cell=f'X{start_row+3}'
    ParameterInOrOut_MergeCell=ParameterInOrOut_Cell+f':Z{start_row+3}'

    ParameterDescription_Cell =f'AA{start_row+3}'
    ParameterDescription_MergeCell=ParameterDescription_Cell+f':AP{start_row+3}'

    ReturnValue_Cell=f'E{start_row+5+parameterColumCorrectionValue}'
    ReturnValue_MergeCell=ReturnValue_Cell+f":H{start_row+6+parameterColumCorrectionValue}"

    ReturnValueDataType_Cell=f'I{start_row+5+parameterColumCorrectionValue}'
    ReturnValueDataType_Value_Cell=f'I{start_row+6+parameterColumCorrectionValue}'
    ReturnValueDataType_MergeCell=ReturnValueDataType_Cell+f':M{start_row+5+parameterColumCorrectionValue}'
    ReturnValueDataType_Value_MergeCell=ReturnValueDataType_Value_Cell+f':M{start_row+6+parameterColumCorrectionValue}'
    
    ReturnValuePossibleReturnValue_Cell=f'N{start_row+5+parameterColumCorrectionValue}'
    ReturnValuePossibleReturnValue_Value_Cell=f'N{start_row+6+parameterColumCorrectionValue}'
    ReturnValuePossibleReturnValue_MergeCell=ReturnValuePossibleReturnValue_Cell+f':W{start_row+5+parameterColumCorrectionValue}'
    ReturnValuePossibleReturnValue_Value_MergeCell=ReturnValuePossibleReturnValue_Value_Cell+f':W{start_row+6+parameterColumCorrectionValue}'
    
    RetrunValueDescription_Cell=f'X{start_row+5+parameterColumCorrectionValue}'
    RetrunValueDescription_MergeCell=RetrunValueDescription_Cell+f':AP{start_row+5+parameterColumCorrectionValue}'
    RetrunValueDescription_Value_Cell=f'X{start_row+6+parameterColumCorrectionValue}'
    RetrunValueDescription_Value_MergeCell=RetrunValueDescription_Value_Cell+f':AP{start_row+6+parameterColumCorrectionValue}'

    # 클래스이름 폰트 설정
    ws[ClassName_Cell].font = Font(name='Arial', size=20, bold=True)
     #----- 기본 테이블 값 셋팅 ----------
    ws[InterfaceName_Cell] = 'Interface Name'
    ws[Prototype_Cell] = 'Prototype'
    ws[Description_Cell] = 'Description'
    ws[Parameter_Cell] = 'Parameter'
    ws[ParameterDataType_Cell] = 'Data Type'
    ws[ParameterName_Cell] = 'Name'
    ws[ParameterRange_Cell] = 'Range'
    ws[ParameterInOrOut_Cell] = 'IN/OUT'
    ws[ParameterDescription_Cell] = 'Description'
    ws[ReturnValue_Cell]='Return Value'
    ws[ReturnValueDataType_Cell]='Data Type'
    ws[ReturnValuePossibleReturnValue_Cell]='Possible Return Value'
    ws[RetrunValueDescription_Cell]='Description'

    #------------컬럼 보정작업을 하고 병합을하고 병합한 셀에 값을 추가해줘야 오류가 없다 !-----
    merge_cells =[InterfaceName_MergeCell,InterfaceName_Value_MergeCell,Prototype_MergeCell,Prototype_Value_MegrgeCell,
                  Description_MergeCell,Description_Value_MergeCell,Parameter_MergeCell,ParameterDataType_MergeCell,
                  ParameterName_MergeCell,ParameterRange_MergeCell,ParameterInOrOut_MergeCell,ParameterDescription_MergeCell,
                  ReturnValue_MergeCell,ReturnValueDataType_MergeCell,ReturnValuePossibleReturnValue_MergeCell,
                  RetrunValueDescription_MergeCell,ReturnValueDataType_Value_MergeCell,ReturnValuePossibleReturnValue_Value_MergeCell,
                  RetrunValueDescription_Value_MergeCell]
    
    for merge_cell in merge_cells:
        ws.merge_cells(merge_cell)

   
    #---------------------------------------------------------------------------

    #---------------------파싱값 할당--------------------------------------------
    ws[InterfaceName_Value_Cell]= unitNameList[unitNameIndex-1]
    ws[Prototype_Value_Cell]= orginalPrototype   #prototype 이다 !

    if ">" in allParameterValue:
        saveParameter = toCharParamter(prototype=prototype)   
        for index in range(0,len(saveParameter)):  # 저장된 파라미터 갯수만큼 돌아가지 
            parameterDataType=''
            for combindParameterDataTypeIndex in range(0,len(saveParameter[index].split(" "))-1): # 목적 파라미터 이름 빼고 앞에 있는걸 획득하기 위함 ; 보통 0 1 로 끝나겠지
                parameterDataType+=saveParameter[index].split(" ")[combindParameterDataTypeIndex]  # 이렇게 해주면 되자나

            #index수만큼 행을 늘린다 이게 왜필요? 파라미터 수에 따라서 병합 해줘야하기때문에 이렇게 5개만들어야하니까 이것도 for문 도릴ㄹ가 ?
           
            # 자.. 좌표로 셀을 병합하고싶다 강렬하게...일단 여긴 하지말자 별로 안나오자나 그럼 밑에서 한번 만들고 실험 .
            correctionColum =Parameter_Element_Value_Row+index       
            ws.merge_cells(start_row=correctionColum, 
                        start_column=ParameterDataType_Value_Merge_Start_Culum, 
                        end_row=correctionColum, 
                        end_column=ParameterDataType_Value_Merge_End_Culum)
            ws.merge_cells(start_row=correctionColum, 
                        start_column=ParameterName_Value_Start_Culum, 
                        end_row=correctionColum, 
                        end_column=ParameterName_Value_End_Culum)         
            ws.merge_cells(start_row=correctionColum, 
                        start_column=ParameterRange_Value_Start_Culum, 
                        end_row=correctionColum, 
                        end_column=ParameterRange_Value_End_Culum)        
            ws.merge_cells(start_row=correctionColum, 
                        start_column=ParameterInOrOut_Value_Start_Culum, 
                        end_row=correctionColum, 
                        end_column=ParameterInOrOut_Value_End_Culum)
            ws.merge_cells(start_row=correctionColum, 
                        start_column=ParameterDescription_Value_Start_Culum, 
                        end_row=correctionColum, 
                        end_column=ParameterDescription_Value_End_Culum)    
       
            ws.cell(Parameter_Element_Value_Row,ParameterDataType_Value_Culum).value = parameterDataType # 파라미터 데이터 타입 위에서 index더해주니까 여기선 적용된 상태니 상관 없지
            ws.cell(Parameter_Element_Value_Row,ParameterName_Value_Culum).value = saveParameter[index].split(" ")[len(saveParameter[index].split(" "))-1] # 파라미터 이름
            ws.cell(Parameter_Element_Value_Row,ParameterInOrOut_Value_Culum).value = "IN"

            ws.cell(correctionColum,ParameterDataType_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
            ws.cell(correctionColum,ParameterName_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
            ws.cell(correctionColum,ParameterInOrOut_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    else :
        for param in parameterList:
            if not param:
                #여긴 가장 실험하기 쉬운 곳 
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterDataType_Value_Merge_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterDataType_Value_Merge_End_Culum)
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterName_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterName_Value_End_Culum)         
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterRange_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterRange_Value_End_Culum)        
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterInOrOut_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterInOrOut_Value_End_Culum)
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterDescription_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterDescription_Value_End_Culum)                
                
                
                ws.cell(Parameter_Element_Value_Row,ParameterDataType_Value_Culum).value ='-'
                ws.cell(Parameter_Element_Value_Row,ParameterName_Value_Culum).value ='-'
                ws.cell(Parameter_Element_Value_Row,ParameterRange_Value_Culum).value ='-'
                ws.cell(Parameter_Element_Value_Row,ParameterInOrOut_Value_Culum).value ='-'
                ws.cell(Parameter_Element_Value_Row,ParameterDescription_Value_Culum).value  ='-'

                ws.cell(Parameter_Element_Value_Row,ParameterDataType_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
                ws.cell(Parameter_Element_Value_Row,ParameterName_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                ws.cell(Parameter_Element_Value_Row,ParameterInOrOut_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                ws.cell(Parameter_Element_Value_Row,ParameterRange_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                ws.cell(Parameter_Element_Value_Row,ParameterDescription_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
            elif len(parameterList) ==1:
                parameter = allParameterValue.split(' ')
                dataType=''
                for index in range(0,len(parameter)-1):
                    dataType+=parameter[index]+" "

                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterDataType_Value_Merge_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterDataType_Value_Merge_End_Culum)
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterName_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterName_Value_End_Culum)         
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterRange_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterRange_Value_End_Culum)        
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterInOrOut_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterInOrOut_Value_End_Culum)
                ws.merge_cells(start_row=Parameter_Element_Value_Row, 
                               start_column=ParameterDescription_Value_Start_Culum, 
                               end_row=Parameter_Element_Value_Row, 
                               end_column=ParameterDescription_Value_End_Culum)              
       
                ws.cell(Parameter_Element_Value_Row,ParameterDataType_Value_Culum).value  = dataType # 파라미터 데이터 타입
                ws.cell(Parameter_Element_Value_Row,ParameterName_Value_Culum).value= parameter[len(parameter)-1] # 파라미터 이름
                ws.cell(Parameter_Element_Value_Row,ParameterInOrOut_Value_Culum).value = "IN"

                ws.cell(Parameter_Element_Value_Row,ParameterDataType_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
                ws.cell(Parameter_Element_Value_Row,ParameterName_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                ws.cell(Parameter_Element_Value_Row,ParameterInOrOut_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 

            else: # 이제 파라미터 값이 여러개 오는 짜증나는 경우다 // 일단 여러줄 만드는거 성공했자나 ;; 이제 값만 잘 넣어주면된다...
                # for문 이용해서 행값 조절 좀 해줘야 할듯 ?
                for parameterColum in range(0,parameterColumCorrectionValue+1):
                    parameter = parameterList[parameterColum].split(' ')
                    dataType=''
                    #Parameter_Element_Value_Merge_Row+parameterColum
                    for index in range(0,len(parameter)-1):
                        dataType+=parameter[index]+" "
                    correctionColum =Parameter_Element_Value_Row+parameterColum       
                    ws.merge_cells(start_row=correctionColum, 
                                start_column=ParameterDataType_Value_Merge_Start_Culum, 
                                end_row=correctionColum, 
                                end_column=ParameterDataType_Value_Merge_End_Culum)
                    ws.merge_cells(start_row=correctionColum, 
                                start_column=ParameterName_Value_Start_Culum, 
                                end_row=correctionColum, 
                                end_column=ParameterName_Value_End_Culum)         
                    ws.merge_cells(start_row=correctionColum, 
                                start_column=ParameterRange_Value_Start_Culum, 
                                end_row=correctionColum, 
                                end_column=ParameterRange_Value_End_Culum)        
                    ws.merge_cells(start_row=correctionColum, 
                                start_column=ParameterInOrOut_Value_Start_Culum, 
                                end_row=correctionColum, 
                                end_column=ParameterInOrOut_Value_End_Culum)
                    ws.merge_cells(start_row=correctionColum, 
                                start_column=ParameterDescription_Value_Start_Culum, 
                                end_row=correctionColum, 
                                end_column=ParameterDescription_Value_End_Culum)                           
        
                    ws.cell(correctionColum,ParameterDataType_Value_Culum).value = dataType # 파라미터 데이터 타입
                    ws.cell(correctionColum,ParameterName_Value_Culum).value = parameter[len(parameter)-1] # 파라미터 이름
                    ws.cell(correctionColum,ParameterInOrOut_Value_Culum).value = "IN"

                    ws.cell(correctionColum,ParameterDataType_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)    
                    ws.cell(correctionColum,ParameterName_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
                    ws.cell(correctionColum,ParameterInOrOut_Value_Culum).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 

    # ----------------적절한 반환형을 주입시키는 분기------------------------------
    if unitNameIndex == 2:
        ws[ReturnValueDataType_Value_Cell] = "-"   
        ws[ReturnValuePossibleReturnValue_Value_Cell] = "-"
        ws[RetrunValueDescription_Value_Cell] = "-"
    elif unitNameIndex == 3:
        if unitNameList[unitNameIndex-2] == "void" or unitNameList[unitNameIndex-2] == "virtual" :
            ws[ReturnValueDataType_Value_Cell] = "-"   
            ws[ReturnValuePossibleReturnValue_Value_Cell] = "-"
            ws[RetrunValueDescription_Value_Cell] = "-"
        else:
            ws[ReturnValueDataType_Value_Cell] = unitNameList[unitNameIndex-2] # 반환형!
    elif unitNameIndex > 3:
        if "void" in orginalPrototype:
            ws[ReturnValueDataType_Value_Cell] = "-"   
            ws[ReturnValuePossibleReturnValue_Value_Cell] = "-"
            ws[RetrunValueDescription_Value_Cell] = "-"
        else:    
            returnValue=''
            for unitIndex in range(1,unitNameIndex-1):
                returnValue +=unitNameList[unitIndex]+" "
            ws[ReturnValueDataType_Value_Cell]=returnValue


    #---------------------------------------------------------------------------
    #---------------------------------------------------------------------------


    # 열 크기 지정
    for col in range(5,43):
        ws.column_dimensions[get_column_letter(col)].width = 8.1 * 1.09


    
    # 색 및 글자 폰트
    col_color = [InterfaceName_Cell,Prototype_Cell,Description_Cell,Parameter_Cell,ReturnValue_Cell,
                 ParameterDataType_Cell,ParameterName_Cell,ParameterRange_Cell,ParameterInOrOut_Cell,ParameterDescription_Cell,
                 ReturnValueDataType_Cell,ReturnValuePossibleReturnValue_Cell,RetrunValueDescription_Cell]
    
    for cell in col_color:
        ws[cell].fill = PatternFill(patternType="solid", fgColor='BFBFBF')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        if cell == f'B{start_row}':
            ws[cell].font = Font(name='Arial', size=10, bold=True)
        else:
            ws[cell].font = Font(name='Arial', size=10)
    
    # 추가 정렬작업
    col_center =[ReturnValueDataType_Value_Cell,ReturnValuePossibleReturnValue_Value_Cell,RetrunValueDescription_Value_Cell]
    for cell in col_center:
         ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
 
    #표 테두리 작업
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for rows in ws.iter_rows(min_row=start_row, max_row=start_row+parameterColumCorrectionValue + 6, min_col=5, max_col=42):
        for cell in rows:
            cell.border = border


def create_class_name(ws,className,checkInterfaceClass):  
   ws.append([None])

   checkName=checkInterfaceClass.split(".")[0].strip() 

   if "::" in className:
    dividedClassName =className.split(":")
    exsistBlankClassName=dividedClassName[len(dividedClassName)-1].split(" ")
    realClassName =exsistBlankClassName[0]
    if "인터페이스" in exsistBlankClassName[1]:
        if realClassName==checkName:            
            interfaceName="(Interface) " + exsistBlankClassName[0]         
            ws.append({'A':interfaceName})
        else:
            ws.append({'A':realClassName})
    else:#클래스겠지 
        if realClassName==checkName:  
          ws.append({'A':realClassName})
        else:
            innerClassName=checkName+" :: "+realClassName
            ws.append({'A':innerClassName})    
   else:
    dividedClassName =className.split(".")
    exsistBlankClassName=dividedClassName[len(dividedClassName)-1].split(" ")
    realClassName =exsistBlankClassName[0]
    if "인터페이스" in exsistBlankClassName[1]:
        if realClassName==checkName:            
            interfaceName="(Interface) " + exsistBlankClassName[0]         
            ws.append({'A':interfaceName})
        else:
             ws.append({'A':realClassName})  
    else:
        if realClassName==checkName:  
            ws.append({'A':realClassName})
        else:
            innerClassName=checkName+" :: "+realClassName
            ws.append({'A':innerClassName})  

def getFileNameInDrectory(output_file_name):
    if "\\" in output_file_name or "/" in output_file_name:
        if "\\" in output_file_name:
            split_output_file_name = output_file_name.split("\\")
            return split_output_file_name[len(split_output_file_name)-1]
        else:
            split_output_file_name = output_file_name.split("/")   
            return split_output_file_name[len(split_output_file_name)-1]
    else:
        return output_file_name

def delete_file(file_path):
    try:
        shutil.rmtree(file_path)
        print(f"파일 {file_path} 삭제 성공")
    except OSError as e:
        print("Error: %s : %s" % (file_path, e.strerror))        

def set_write_permission(file_path):
    # 쓰기 권한을 부여할 권한 값 (octal notation)
    write_permission = 0o666
    try:
        # 파일에 대한 쓰기 권한 설정
        os.chmod(file_path, write_permission)
        print("파일 쓰기 권한이 설정되었습니다.")
    except OSError as e:
        print(f"파일 권한 변경 실패: {e}")

def create_html_file(dir_path,isJavafile,deepParsing,output_file_name):
    try:
            isJavafileYesOrNo=''
            isCppFileYesOrNo=''
            if isJavafile.lower() in ["y","yes"]:
                isJavafileYesOrNo = "YES"
                isCppFileYesOrNo="NO"
            elif isJavafile.lower() in ["n","no"]:
                isJavafileYesOrNo = "NO"
                isCppFileYesOrNo="YES"
            else:
                raise Exception("Y or N 중에 하나만 입력해 주세요") 
    except Exception as e:
        print(e)
    try:          
            deepParsingYesOrNo=''
            if deepParsing.lower() in ["y","yes"]:
                deepParsingYesOrNo="YES"
            elif deepParsing.lower() in ["n","no"]:
                deepParsingYesOrNo="NO" 
            else:     
                raise Exception("Y or N 중에 하나만 입력해 주세요")
    except Exception as e:
        print(e)    

    # output_file_name / or \ 있다면 파일경로가 포함된 저장폴더 이름이니까 스플릿하고 아니면 그대로 반환     
    file_end_name=getFileNameInDrectory(output_file_name)

    #만약에 동일 이름의 파일이 존재하면 파일을 삭제하고 다시만드는 과정을 추가 
    real_file_path =f'{dir_path}/HTML_{file_end_name}'
    isSameFileExist=os.path.exists(real_file_path)    
    if isSameFileExist:
        print("기존에 존재하는 파일을 삭제하고 다시 생성합니다")
        set_write_permission(real_file_path)
        delete_file(real_file_path)
    
    # 설정 값을 저장할 딕셔너리 생성
    config = {}

    # 프로젝트 이름
    config['PROJECT_NAME'] = {}
    config['PROJECT_NAME']['PROJECT_NAME'] = 'made by antony'

    config['OUTPUT_DIRECTORY'] = {}
    # 인풋값으로 들어올것.. 이 경로는 html이 모일장소니 따로 설정해주거나 파싱해야겠지?
    # 파싱하는게 안전하다고 생각이 들긴해 그럼 소스폴더 파일에 경로에 따로 하나 만들어주자 그게 더 좋을듯 해
    
    config['OUTPUT_DIRECTORY']['OUTPUT_DIRECTORY'] = real_file_path

    # 언어설정
    config['OUTPUT_LANGUAGE'] = {}
    config['OUTPUT_LANGUAGE']['OUTPUT_LANGUAGE'] = 'Korean'

    # JAVA 파일이라면 
    config['OPTIMIZE_OUTPUT_JAVA'] = {}
    config['OPTIMIZE_OUTPUT_JAVA']['OPTIMIZE_OUTPUT_JAVA'] = isJavafileYesOrNo
    # CPP 파일일 경우 
    config['CPP_CLI_SUPPORT'] = {}
    config['CPP_CLI_SUPPORT']['CPP_CLI_SUPPORT'] = isCppFileYesOrNo

    # INPUT 경로 즉 .. 소스파일이 위치하는 경로인데 컴포먼트별로 입력 받도록 해야할듯
    config['INPUT'] = {}
    config['INPUT']['INPUT'] = dir_path

    # 뭐 깊은 탐색인가 그거 하는거인듯
    config['RECURSIVE'] = {}
    config['RECURSIVE']['RECURSIVE'] = deepParsingYesOrNo

    #extract 설정
    config['EXTRACT_ALL'] = {}
    config['EXTRACT_ALL']['EXTRACT_ALL'] = 'YES'
    config['EXTRACT_PRIVATE'] = {}
    config['EXTRACT_PRIVATE']['EXTRACT_PRIVATE'] = 'YES'
    config['EXTRACT_PACKAGE'] = {}
    config['EXTRACT_PACKAGE']['EXTRACT_PACKAGE'] = 'YES'
    config['EXTRACT_STATIC'] = {}
    config['EXTRACT_STATIC']['EXTRACT_STATIC'] = 'YES'

    #class_graph 설정인듯
    config['CLASS_GRAPH'] = {}
    config['CLASS_GRAPH']['CLASS_GRAPH'] = 'NO'

    # 설정 파일로 저장
    file_path = f'{dir_path}/configuration'
    with open(file_path, 'w') as configfile:
        for section, options in config.items():
            #configfile.write(f'[{section}]\n')
            for option, value in options.items():
                configfile.write(f'{option} = {value}\n')
            configfile.write('\n')

    # cmd 로 저장된 파일 실행 
    terminnal_command ='doxygen configuration'
    os.chdir(dir_path)
    os.system(terminnal_command)


        
def main(file_path, output_file_name,isJavafile,deepParsing):

    create_html_file(file_path,isJavafile,deepParsing,output_file_name)

    print("\n\n*******************엑셀파일이 생성될 때까지 기다려주세요*******************\n\n")
    table_texts = []
    titles = []
    source_files=[]

    file_end_name=getFileNameInDrectory(output_file_name) 
    html_file_path=f'{file_path}/HTML_{file_end_name}'
    for (root, directories, files) in os.walk(html_file_path):
        for file in files:
            if '.html' in file:
                html_file_path = os.path.join(root, file)
                # print(file_path)
                tables = find_tables_with_methods(html_file_path)
                title_text = find_title_text(html_file_path)
                source_file = find_source_file(html_file_path)
                if tables and title_text: # -mothod 테이블이 없으면 그 파일은 append 안됨 !
                    table_texts.append([table.get_text() for table in tables])
                    titles.append(title_text)
                    source_files.append(source_file)    
    # 새 워크북 생성 및 활성 시트 선택
    wb = Workbook()
    wb_ADS = Workbook()
    
    count = 0
    count_ADS =0
    startColumCorrectionValue = 0
    startColumCorrectionValue_ADS =0
    countClassName = 0
    ws = wb.active
    ws_ADS = wb_ADS.active
    for data in table_texts:
        className = titles[countClassName]
        checkInterfaceClass=source_files[countClassName]
        create_class_name(ws, className=className,checkInterfaceClass=checkInterfaceClass)
        create_class_name(ws_ADS, className=className,checkInterfaceClass=checkInterfaceClass)
        countClassName += 1
        for splitData in data:
                accessModifier =''
                for creTable in splitData.split("\n"):
                    creTable = creTable.strip()  # 앞뒤 공백 제거
                    creTable = creTable.strip().replace('\xa0', ' ') # 짜증나는 녀석 제거
                    if not creTable:  
                        continue
                    if "멤버" in creTable:
                        accessModifier = creTable.split(' ')[0].lower()
                        if accessModifier =="정적":
                            accessModifier = creTable.split(' ')[1].lower()
                        continue
                    if ")" not in creTable:
                        continue
                    if "(으)로부터" in accessModifier:
                        continue
                    if accessModifier or not accessModifier: # 접근제어자가 존재하지 않을 수도 있다는 사실이 존재함 ............
                        prototype = accessModifier + " " + creTable 
                        start_row = 2 + count * 14 +  startColumCorrectionValue+countClassName
                        saveParameter = toCharParamter(prototype=prototype)
                        parameterColumCorrectionValue=len(saveParameter)-1  
                        create_table_DDS(ws, start_row=start_row, prototype=prototype,parameterColumCorrectionValue=parameterColumCorrectionValue)#보정값2이 들어감 들어간 보정값은 prameter 이후의 셀들에 더해지고)
                        count += 1
                        startColumCorrectionValue+=parameterColumCorrectionValue #보정값2을 더해줌 들어간 보정값1에.. start_rowdpeh 적용 될수 있게

                    if "private" not in accessModifier:
                        prototype = accessModifier + " " + creTable 
                        start_row = 2 + count_ADS * 8 +  startColumCorrectionValue_ADS+countClassName
                        saveParameter = toCharParamter(prototype=prototype)
                        parameterColumCorrectionValue=len(saveParameter)-1  
                        create_table_ADS(ws_ADS, start_row=start_row, prototype=prototype,parameterColumCorrectionValue=parameterColumCorrectionValue)#보정값2이 들어감 들어간 보정값은 prameter 이후의 셀들에 더해지고)
                        count_ADS += 1
                        startColumCorrectionValue_ADS+=parameterColumCorrectionValue #보정값2을 더해줌 들어간 보정값1에.. start_rowdpeh 적용 될수 있게 

    root= Tk()
    root.withdraw()

    msg.showinfo('made by antony', "클래스 총 갯수 : " +str(len(titles)) +"\n"+
                 "DDS 테이블 생성 갯수 :  "+str(count) +"\n"+
                 "ADS 테이블 생성 갯수 : "+str(count_ADS))

    print("================================= 클래스 총 갯수       : "+str(len(titles))+"     =================================")
    print("================================= DDS 테이블 생성 갯수 : "+str(count)+"     =================================")
    print("================================= ADS 테이블 생성 갯수 : "+str(count_ADS)+"     =================================")
    print("                                    made by antony                                         ")
    os.system("pause")

    # print(len(table_texts))
    # print(len(titles))
    # print(len(source_files))

    # for index in range(0,len(titles)):
    #     print(f"\n{titles[index]}         vs        "+f'{source_files[index]}\n')

    wb.save(f'{output_file_name}_DDS.xlsx')
    wb_ADS.save(f'{output_file_name}_ADS.xlsx')


if __name__ == "__main__":
    print("*************파일경로에 띄어쓰기나 한글이 적혀있으면 경로를 인식을 못합니다*************")
    print("파일 경로를 입력하세요:")
    file_path = input()


    print("********************************************************************************************")
    print("        저장할 경로를 명시하지 않는 다면 소스파일이 있는 디렉토리에 생성됩니다   ")
    print("        생성될 엑셀 파일의 [저장 파일 이름] 또는 [저장경로\저장파일 이름]을 입력하세요 :    ") 
    print("********************************************************************************************")
    output_file_name = input()

    countIsJavaInputTry = 0
    while(True):
        print("자바파일 인가요?(Y/N)")
        isJavafile=input()
        try:
            if isJavafile.lower() in ["y", "n","yes","no"]:
                break
            elif countIsJavaInputTry==5:
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                os.system("pause")
                sys.exit(1)                     
            else:    
                raise Exception("\n\nY or N 을 입력해주세요")        
        except Exception as e:
            countIsJavaInputTry+=1
            print(e)    
    
    countDeepParsingInputTry =0
    while(True):
        print("깊은 탐색을 해서 하위 폴더의 html까지 파싱하고 싶으신가요?(Y/N)")
        deepParsing=input()
        try:
            if deepParsing.lower() in ["y", "n","yes","no"]:
                break
            elif countDeepParsingInputTry==5:
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                print("사용법을 숙지하시고 다시 시도하시길 바랍니다.")
                os.system("pause")
                sys.exit(1)                   
            else:
                raise Exception("\n\nY or N 을 입력해주세요 - antony")
        except Exception as e:
            countDeepParsingInputTry+=1
            print(e)
           
    main(file_path, output_file_name,isJavafile,deepParsing)

#1. 엑셀 저장 경로를 추가해주자(해결)    
#2. html 긁어오는건 생성되는 html파일기준으로 하기 ..(해결)
#3. 생성할때 동일 HTML 파일 이름 존재하면 삭제하고 다시 만드는것..!(해결 파일 삭제가 아닌 폴더삭제임.!)
#4. ADS 추가하기 (해결!)
    #4-1. ADS private 제외 하고 생성(해결)
    #4-2. 줄 맞춤 (해결)
    #4-3. 양식 생성 (해결)
#번외. 아이콘넣기 (해결)
#번외. Y N 입력시 예외 메세지 설정하기(해결!이야 이거 바로 input들어올때 무한루프 예외 처리하는걸로 . .)
    # 5번 실패시 프로세스 pasue하고 종료 해야할듯 함..!(해결)
#번외 print 말고 메세지 박스 이용하기(해결)
    # 에러나 결과에대해서 ..~(해결)


#인터페이스 구분기능이 있었으며 좋겠다라는 생각을 하는중
 # 현재상황 인터페이스
    # 인터페이스는 두가지 종류가 존재해 
        # 1. 진짜 파일이 인터페이스파일 -> 클래스와 구분이 안되니 구분필요 (Interface) 인터페이스명 (해결)
        # 2. 클래스 안에 정의한 인터페이스or 인터페이스->  인터페이스:: 메소드명 이건 좀 힘드네 ;;;
            #2-1 클래스의 경우는 아마... 해당 이너클래스 :: 클래스명 이걸로 했음.. 이유는 여러 클래스에서 중복되게 정의해놓아서 이건 클래스 OK!
        # 3... ADS의 경우는,, 이너 클래스 인터페이스 안한듯;
            #3-1 이경우에는 ... 안맞으면 그냥 패스하는걸로;;; ;
#예외처리 생각하기
#번외 GUI